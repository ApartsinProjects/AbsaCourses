"""Score the human annotation study (N1/H1): human-vs-declared, inter-rater
agreement, and human-vs-audit agreement on the synthetic corpus.

Pattern 5 (per-(item, aspect): present? + sentiment). Three raters, blind to
declared labels. Joins:
  rater_res.xlsx  (item_id -> present Yes/No, sentiment)
  annotation_key.csv (item_id -> sample_id, aspect, declared_sentiment, audit_row_score)
  at_scale audit responses.jsonl (row_id -> per-aspect supported / sentiment_match)
"""
import csv, json, itertools, statistics as st
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[2]
HR = Path(__file__).resolve().parent
KEY = ROOT / "paper/human_annotation/annotation_key.csv"
AUDIT = ROOT / "paper/faithfulness_audit/at_scale_gpt-4.1-mini_responses.jsonl"
RATERS = ["rater1", "rater2", "rater3"]


def cohen_kappa(a, b):
    """Cohen's kappa for two equal-length label lists."""
    n = len(a)
    labels = sorted(set(a) | set(b))
    po = sum(1 for x, y in zip(a, b) if x == y) / n
    pe = sum((a.count(l) / n) * (b.count(l) / n) for l in labels)
    return (po - pe) / (1 - pe) if pe != 1 else 1.0, po


def fleiss_kappa(rows):
    """rows: list of dict{label: count}; each row same total raters."""
    N = len(rows)
    n = sum(rows[0].values())
    cats = sorted({c for r in rows for c in r})
    p = {c: sum(r.get(c, 0) for r in rows) / (N * n) for c in cats}
    Pe = sum(v * v for v in p.values())
    Pbar = sum((sum(r.get(c, 0) ** 2 for c in cats) - n) / (n * (n - 1)) for r in rows) / N
    return (Pbar - Pe) / (1 - Pe) if Pe != 1 else 1.0


def load_rater(name):
    wb = openpyxl.load_workbook(HR / f"annotation_{name}_res.xlsx")
    ws = wb["annotation"] if "annotation" in wb.sheetnames else wb.active
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        iid, _text, _asp, present, sent = row[0], row[1], row[2], row[3], (row[4] if len(row) > 4 else None)
        pres = str(present).strip().lower() if present else None
        out[iid] = {"present": "yes" if pres == "yes" else "no",
                    "sentiment": (str(sent).strip().lower() if sent else None)}
    return out


def main():
    key = {r["item_id"]: r for r in csv.DictReader(open(KEY, encoding="utf-8"))}
    raters = {n: load_rater(n) for n in RATERS}
    items = [i for i in key if all(i in raters[n] for n in RATERS)]

    # audit per-aspect judgments keyed by (sample_id, aspect_normalized)
    audit = {}
    for line in open(AUDIT, encoding="utf-8"):
        rec = json.loads(line)
        rid = str(rec["row_id"])
        for a in rec.get("judged", {}).get("aspects", []):
            audit[(rid, a["aspect"].replace("_", " "))] = {
                "supported": bool(a.get("supported")),
                "sentiment_match": bool(a.get("sentiment_match"))}

    res = {"n_items": len(items), "n_raters": len(RATERS)}

    # ---- 1. Human-vs-declared -------------------------------------------------
    # Every item is a DECLARED aspect, so declared-present = yes by construction.
    hvd = {}
    for n in RATERS:
        pres = [1 if raters[n][i]["present"] == "yes" else 0 for i in items]
        conf = sum(pres) / len(pres)  # human confirms the declared aspect is present
        # sentiment-match among items this rater marked present with a declared sentiment
        sm = [(raters[n][i]["sentiment"] == key[i]["declared_sentiment"])
              for i in items if raters[n][i]["present"] == "yes"
              and key[i]["declared_sentiment"] and raters[n][i]["sentiment"]]
        hvd[n] = {"presence_confirm_rate": round(conf, 3),
                  "sentiment_match_rate_when_present": round(sum(sm) / len(sm), 3) if sm else None,
                  "n_present": sum(pres)}
    # majority vote
    def maj_present(i):
        return sum(1 for n in RATERS if raters[n][i]["present"] == "yes") >= 2
    maj_conf = sum(1 for i in items if maj_present(i)) / len(items)
    res["human_vs_declared"] = {"per_rater": hvd,
                                "majority_presence_confirm_rate": round(maj_conf, 3)}

    # ---- 2. Inter-rater agreement --------------------------------------------
    pres_lists = {n: [raters[n][i]["present"] for i in items] for n in RATERS}
    pair_k = {}
    for a, b in itertools.combinations(RATERS, 2):
        k, po = cohen_kappa(pres_lists[a], pres_lists[b])
        pair_k[f"{a}-{b}"] = {"kappa": round(k, 3), "percent_agree": round(po, 3)}
    fleiss_rows = [{"yes": sum(1 for n in RATERS if raters[n][i]["present"] == "yes"),
                    "no": sum(1 for n in RATERS if raters[n][i]["present"] == "no")} for i in items]
    # sentiment agreement among items all 3 marked present
    allp = [i for i in items if all(raters[n][i]["present"] == "yes" for n in RATERS)]
    sent_all_agree = sum(1 for i in allp if len({raters[n][i]["sentiment"] for n in RATERS}) == 1) / len(allp) if allp else None
    res["inter_rater"] = {"presence_cohen_pairwise": pair_k,
                          "presence_fleiss_kappa": round(fleiss_kappa(fleiss_rows), 3),
                          "sentiment_all3_present_exact_agree": round(sent_all_agree, 3) if sent_all_agree else None,
                          "n_all3_present": len(allp)}

    # ---- 3. Human-vs-audit (aspect level) ------------------------------------
    matched = [i for i in items if (key[i]["sample_id"], key[i]["aspect"]) in audit]
    h_pres = [1 if maj_present(i) else 0 for i in matched]
    a_pres = [1 if audit[(key[i]["sample_id"], key[i]["aspect"])]["supported"] else 0 for i in matched]
    kp, pop = cohen_kappa(h_pres, a_pres)
    # sentiment: among items where BOTH human-majority present and audit supported
    both = [i for i in matched if maj_present(i) and audit[(key[i]["sample_id"], key[i]["aspect"])]["supported"]]
    # human-majority sentiment
    def maj_sent(i):
        votes = [raters[n][i]["sentiment"] for n in RATERS if raters[n][i]["present"] == "yes" and raters[n][i]["sentiment"]]
        return st.mode(votes) if votes else None
    sent_agree = [(maj_sent(i) == key[i]["declared_sentiment"]) for i in both if key[i]["declared_sentiment"]]
    res["human_vs_audit"] = {
        "n_matched_aspects": len(matched),
        "presence_cohen_kappa": round(kp, 3), "presence_percent_agree": round(pop, 3),
        "human_vs_declared_sentiment_when_both_present": round(sum(sent_agree) / len(sent_agree), 3) if sent_agree else None,
        "n_both_present": len(both)}

    # ---- 4. Audit validation: human confirmation by audit_row_score quartile --
    scored = sorted(items, key=lambda i: float(key[i]["audit_row_score"]))
    q = len(scored) // 4
    quarts = {"Q1_low": scored[:q], "Q2": scored[q:2 * q], "Q3": scored[2 * q:3 * q], "Q4_high": scored[3 * q:]}
    strat = {}
    for name, ql in quarts.items():
        conf = sum(1 for i in ql if maj_present(i)) / len(ql)
        avg = st.mean(float(key[i]["audit_row_score"]) for i in ql)
        strat[name] = {"n": len(ql), "audit_row_score_mean": round(avg, 3),
                       "human_majority_confirm_rate": round(conf, 3)}
    res["audit_validation_by_stratum"] = strat

    out = HR / "human_study_scores.json"
    json.dump(res, open(out, "w"), indent=2)
    print(json.dumps(res, indent=2))
    print(f"\nwrote {out}")


if __name__ == "__main__":
    main()
