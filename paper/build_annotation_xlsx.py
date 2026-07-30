"""Build per-rater Excel annotation workbooks (dropdown validation) for the
human validation of the synthetic corpus (reviewer ask E1/N1/H1).

Design: ~300 synthetic reviews stratified by audit-score quartile (faithful and
low-scored rows both covered), exploded to one row per declared aspect (~600
decisions). Three raters, item order shuffled per rater, BLIND to the declared
sentiment. Dropdowns: present (Yes/No) and sentiment (positive/neutral/negative/
unsure). A key file (declared labels + audit score) is written separately for
scoring and is NOT given to raters.
"""
import csv
import json
import random
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
CORPUS = ROOT / "paper/generated_datasets/batch_69cc15c483488190941478aa4e3a976d_generated_reviews.jsonl"
SCORES = ROOT / "paper/faithfulness_audit/at_scale_gpt-4.1-mini_per_row_scores.csv"
OUTDIR = ROOT / "paper/human_annotation"
N_REVIEWS = 300
RATERS = ["rater1", "rater2", "rater3"]
SEED = 20260705


def load_corpus():
    rows = {}
    for l in open(CORPUS, encoding="utf-8"):
        r = json.loads(l)
        rows[str(r.get("sample_id"))] = {"text": r["text"], "aspects": r.get("aspects") or {}}
    return rows


def load_scores():
    s = {}
    for r in csv.DictReader(open(SCORES, encoding="utf-8")):
        s[str(r["row_id"])] = float(r["row_score"])
    return s


def stratified_sample(corpus, scores, n, rng):
    ids = [i for i in corpus if i in scores and corpus[i]["aspects"]]
    ids.sort(key=lambda i: scores[i])
    q = len(ids) // 4
    quartiles = [ids[:q], ids[q:2 * q], ids[2 * q:3 * q], ids[3 * q:]]
    per_q = n // 4
    picked = []
    for ql in quartiles:
        picked += rng.sample(ql, min(per_q, len(ql)))
    return picked


def build_items(corpus, scores, sampled):
    """One decision row per (review, declared aspect). Declared sentiment kept in key only."""
    items = []
    for sid in sampled:
        rec = corpus[sid]
        for aspect, declared_sent in rec["aspects"].items():
            items.append({
                "item_id": f"r{sid}_{aspect}",
                "sample_id": sid,
                "review_text": rec["text"],
                "aspect": aspect.replace("_", " "),
                "declared_sentiment": declared_sent,   # key only, not shown to raters
                "audit_row_score": round(scores[sid], 3),
            })
    return items


def write_rater_workbook(items, path, rater):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "annotation"
    header = ["item_id", "review_text", "aspect_to_judge",
              "aspect_present? (Yes/No)", "sentiment_if_present", "other_aspects_you_notice", "notes"]
    ws.append(header)
    hfill = PatternFill("solid", fgColor="14385C")
    for c in range(1, len(header) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for it in items:
        ws.append([it["item_id"], it["review_text"], it["aspect"], "", "", "", ""])
    # dropdowns
    n = len(items) + 1
    dv_present = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False)
    dv_sent = DataValidation(type="list", formula1='"positive,neutral,negative,unsure"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_present)
    ws.add_data_validation(dv_sent)
    dv_present.add(f"D2:D{n}")
    dv_sent.add(f"E2:E{n}")
    # widths + wrap
    widths = {"A": 20, "B": 82, "C": 22, "D": 20, "E": 20, "F": 26, "G": 26}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, max_row=n):
        for c in row:
            c.alignment = Alignment(wrap_text=(c.column_letter == "B"), vertical="top")
    ws.freeze_panes = "A2"
    # instructions sheet
    ins = wb.create_sheet("instructions", 0)
    lines = [
        f"Human annotation - {rater}",
        "",
        "For each row you are shown a student course review and ONE aspect to judge.",
        "1) aspect_present?  -> Yes if the review clearly expresses that aspect, else No.",
        "2) sentiment_if_present -> if present, the student's sentiment toward that aspect",
        "   (positive / neutral / negative). Use 'unsure' only if genuinely ambiguous.",
        "   Leave blank if aspect_present? is No.",
        "3) other_aspects_you_notice -> optional: list any aspect clearly present that is NOT the one asked.",
        "4) notes -> optional.",
        "",
        "Judge only from the review text. You are NOT shown the intended label; that is deliberate.",
        "Use the dropdowns in columns D and E. Work top to bottom; the order is shuffled for you.",
    ]
    for i, t in enumerate(lines, 1):
        ins.cell(i, 1, t)
        if i == 1:
            ins.cell(i, 1).font = Font(bold=True, size=13)
    ins.column_dimensions["A"].width = 100
    wb.save(path)


def main():
    OUTDIR.mkdir(parents=True, exist_ok=True)
    corpus, scores = load_corpus(), load_scores()
    rng = random.Random(SEED)
    sampled = stratified_sample(corpus, scores, N_REVIEWS, rng)
    items = build_items(corpus, scores, sampled)
    # key file (declared labels + audit score) for scoring - NOT for raters
    with (OUTDIR / "annotation_key.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["item_id", "sample_id", "aspect", "declared_sentiment", "audit_row_score"])
        w.writeheader()
        for it in items:
            w.writerow({k: it[k] for k in ["item_id", "sample_id", "aspect", "declared_sentiment", "audit_row_score"]})
    # per-rater shuffled workbooks
    for ri, rater in enumerate(RATERS):
        shuffled = items[:]
        random.Random(SEED + 1 + ri).shuffle(shuffled)
        write_rater_workbook(shuffled, OUTDIR / f"annotation_{rater}.xlsx", rater)
    print(f"reviews sampled={len(sampled)}  decision items/rater={len(items)}  raters={len(RATERS)}")
    print(f"quartile coverage (audit row_score): min={min(scores[s] for s in sampled):.2f} max={max(scores[s] for s in sampled):.2f}")
    print(f"wrote {OUTDIR}/annotation_rater[1-3].xlsx + annotation_key.csv")


if __name__ == "__main__":
    main()
